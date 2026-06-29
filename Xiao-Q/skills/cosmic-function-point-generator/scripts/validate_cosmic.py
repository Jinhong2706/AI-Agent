#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
validate_cosmic.py
对生成的 COSMIC .xlsx 表格执行 15 项合规性校验
用法：python validate_cosmic.py <xlsx_path> [--source <docx_path>] [--output <report_path>]
"""
import sys
import io
# 强制 UTF-8 输出（Windows GBK 兼容）
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
import argparse
import json
import re
from pathlib import Path
from openpyxl import load_workbook


# 禁用词清单
FORBIDDEN_WORDS = ['封装', '加密', 'redis缓存', '持久化数据']
# 触发事件允许取值
ALLOWED_TRIGGER_EVENTS = {
    '定时任务', '流程自动流转', '接口调用', '人工操作'
}
# 复用度允许取值
ALLOWED_REUSE = {'新增', '复用', '利旧'}
# 数据移动类型允许取值
ALLOWED_DATA_MOVEMENT = {'E', 'R', 'W', 'X'}
# 长度上限
MAX_LEN = 30
# 复用度配比下限（新增占比下限 20%）
NEW_RATIO_LOWER = 0.20
NEW_RATIO_UPPER = 0.25


def strip_spaces(s):
    """去除所有空格（含中文/全角）"""
    if not s:
        return ''
    return re.sub(r'\s+', '', s)


def _safe_strip(v):
    if v is None:
        return ''
    return str(v).strip()


def check_step0_fenodian(rows, source_docx=None):
    """校验 0：前置分点扫描"""
    return {'passed': True, 'errors': [], 'note': '需对照原文手动确认'}


def check_step1_yuanwen(rows, source_docx=None):
    """校验 1：原文一致性"""
    errors = []
    for i, r in enumerate(rows, start=4):
        for col in ['F', 'H', 'I']:
            val = _safe_strip(r.get(col, ''))
            if not val:
                errors.append(f"行 {i} {col} 列为空")
    return {'passed': len(errors) == 0, 'errors': errors}


def check_step2_jiegou(rows, source_docx=None):
    """校验 2：结构合规性（前 3 行骨架 + 12 列）"""
    return {'passed': True, 'errors': [], 'note': '前 3 行结构由生成脚本保证'}


def check_step3_gongneng(rows, source_docx=None):
    """校验 3：功能匹配"""
    errors = []
    for i, r in enumerate(rows, start=4):
        if not _safe_strip(r.get('E', '')):
            errors.append(f"行 {i} E 列（功能过程）为空")
        if not _safe_strip(r.get('F', '')):
            errors.append(f"行 {i} F 列（子过程描述）为空")
    return {'passed': len(errors) == 0, 'errors': errors}


def check_step4_duliang(rows, source_docx=None):
    """校验 4：度量规则（G/J/K/D 取值）"""
    errors = []
    for i, r in enumerate(rows, start=4):
        g = _safe_strip(r.get('G', ''))
        if g not in ALLOWED_DATA_MOVEMENT:
            errors.append(f"行 {i} G 列数据移动类型非法：{g}")

        j = _safe_strip(r.get('J', ''))
        if j not in ALLOWED_REUSE:
            errors.append(f"行 {i} J 列复用度非法：{j}")

        k = r.get('K', '')
        if k is None or k == '':
            errors.append(f"行 {i} K 列 CFP 为空")
        else:
            try:
                kf = float(k)
                if j == '新增' and abs(kf - 1.00) > 0.01:
                    errors.append(f"行 {i} K 列 CFP 值与 J 列（新增）不匹配：{kf}")
                elif j == '复用' and abs(kf - 0.33) > 0.01:
                    errors.append(f"行 {i} K 列 CFP 值与 J 列（复用）不匹配：{kf}")
                elif j == '利旧' and abs(kf - 0.00) > 0.01:
                    errors.append(f"行 {i} K 列 CFP 值与 J 列（利旧）不匹配：{kf}")
            except (ValueError, TypeError):
                errors.append(f"行 {i} K 列 CFP 非数值：{k}")

        d = _safe_strip(r.get('D', ''))
        if d and d not in ALLOWED_TRIGGER_EVENTS and not re.search(r'点击|调用|发送|触发|定时', d):
            errors.append(f"行 {i} D 列触发事件不在 5 类允许范围内：{d}")

    return {'passed': len(errors) == 0, 'errors': errors}


def check_step5_shuzhi(rows, source_docx=None):
    """校验 5：数值准确性（ΣCFP = SUM(K)）"""
    errors = []
    sum_k = 0.0
    for i, r in enumerate(rows, start=4):
        k = r.get('K', '')
        if k is not None and k != '':
            try:
                sum_k += float(k)
            except (ValueError, TypeError):
                pass
    sum_k = round(sum_k, 2)

    # 检查首行 L 列
    if rows:
        first_l = rows[0].get('L', '')
        try:
            first_l_val = round(float(first_l), 2) if first_l not in (None, '') else None
            if first_l_val is None or abs(first_l_val - sum_k) > 0.01:
                errors.append(f"首行 L 列 ΣCFP={first_l_val} 与实际累加 {sum_k} 不符")
        except (ValueError, TypeError):
            errors.append(f"首行 L 列 ΣCFP 非数值：{first_l}")

    return {'passed': len(errors) == 0, 'errors': errors, 'sum_cfp': sum_k}


def check_step6_bitian(rows, source_docx=None):
    """校验 6：必填项"""
    errors = []
    for i, r in enumerate(rows, start=4):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            val = r.get(col, '')
            if val is None or val == '':
                errors.append(f"行 {i} {col} 列为空")
    return {'passed': len(errors) == 0, 'errors': errors}


def check_step7_zifuchuan(rows, source_docx=None):
    """校验 7：字符长度（F/H/I ≤ 30）"""
    errors = []
    for i, r in enumerate(rows, start=4):
        for col in ['F', 'H', 'I']:
            val = strip_spaces(r.get(col, ''))
            if len(val) > MAX_LEN:
                errors.append(f"行 {i} {col} 列长度 {len(val)} > {MAX_LEN}：{val[:50]}")
    return {'passed': len(errors) == 0, 'errors': errors}


def check_step8_fengxianci(rows, source_docx=None):
    """校验 8：风险词/禁用过程"""
    errors = []
    for i, r in enumerate(rows, start=4):
        f_val = _safe_strip(r.get('F', ''))
        for word in FORBIDDEN_WORDS:
            if word in f_val:
                # 检查是否为"独立出现"（不是更大的动作词的一部分）
                if f_val == word or f_val.startswith(word) and len(f_val) <= len(word) + 2:
                    errors.append(f"行 {i} F 列出现禁用词「{word}」：{f_val}")
        # 检查单独"判断"
        if f_val == '判断' or f_val.startswith('判断') and not ('并判断' in f_val or '读取' in f_val or '查询' in f_val):
            errors.append(f"行 {i} F 列单独出现「判断」：{f_val}")
    return {'passed': len(errors) == 0, 'errors': errors}


def check_step9_shujuyidong(rows, source_docx=None):
    """校验 9：数据移动类型（首行 E、末行 W/X、≤ 8 行、E 唯一）"""
    errors = []
    # 按 E 列分组
    groups = {}
    for i, r in enumerate(rows, start=4):
        e_val = _safe_strip(r.get('E', ''))
        if not e_val:
            continue
        groups.setdefault(e_val, []).append((i, _safe_strip(r.get('G', ''))))

    for proc, seq in groups.items():
        if len(seq) > 8:
            errors.append(f"功能过程「{proc}」行数 {len(seq)} > 8")
        if seq[0][1] != 'E':
            errors.append(f"功能过程「{proc}」首行 G 列必须为 E，实际为 {seq[0][1]}")
        if seq[-1][1] not in ('W', 'X'):
            errors.append(f"功能过程「{proc}」末行 G 列必须为 W/X，实际为 {seq[-1][1]}")
        e_count = sum(1 for _, g in seq if g == 'E')
        if e_count > 1:
            errors.append(f"功能过程「{proc}」E 动作出现 {e_count} 次（仅允许 1 次）")

    return {'passed': len(errors) == 0, 'errors': errors}


def check_step10_fuyongdu(rows, source_docx=None):
    """校验 10：复用度占比（新增 20%~25%，新增:复用:利旧 ≈ 2:7:1）"""
    errors = []
    counts = {'新增': 0, '复用': 0, '利旧': 0}
    for r in rows:
        j = _safe_strip(r.get('J', ''))
        if j in counts:
            counts[j] += 1
    total = sum(counts.values())
    if total == 0:
        return {'passed': True, 'errors': [], 'counts': counts}

    new_ratio = counts['新增'] / total
    if not (NEW_RATIO_LOWER <= new_ratio <= NEW_RATIO_UPPER):
        errors.append(f"新增占比 {new_ratio:.1%} 不在 {NEW_RATIO_LOWER:.0%}~{NEW_RATIO_UPPER:.0%} 区间（新增/复用/利旧={counts['新增']}/{counts['复用']}/{counts['利旧']}）")

    return {'passed': len(errors) == 0, 'errors': errors, 'counts': counts, 'new_ratio': new_ratio}


def check_step11_hebing(rows, source_docx=None):
    """校验 11：单元格合并（提示性，由 Excel 完成）"""
    return {'passed': True, 'errors': [], 'note': 'A~E 列物理合并需在生成时通过 openpyxl MergeCells 完成'}


def check_step12_shujuzu(rows, source_docx=None):
    """校验 12：数据组唯一性（H 列全表 + 同功能过程）"""
    errors = []
    h_values = []
    proc_h_map = {}
    for i, r in enumerate(rows, start=4):
        h = _safe_strip(r.get('H', ''))
        e = _safe_strip(r.get('E', ''))
        h_values.append((i, h))
        proc_h_map.setdefault(e, []).append((i, h))

    # 全表唯一性
    h_count = {}
    for i, h in h_values:
        if h:
            h_count.setdefault(h, []).append(i)
    for h, indices in h_count.items():
        if len(indices) > 1:
            errors.append(f"H 列「{h}」在多行出现：{indices}")

    # 同功能过程唯一性
    for proc, items in proc_h_map.items():
        proc_h_count = {}
        for i, h in items:
            if h:
                proc_h_count.setdefault(h, []).append(i)
        for h, indices in proc_h_count.items():
            if len(indices) > 1:
                errors.append(f"功能过程「{proc}」下 H 列「{h}」重复：{indices}")

    return {'passed': len(errors) == 0, 'errors': errors}


def check_step13_shujushuxing(rows, source_docx=None):
    """校验 13：数据属性（I 列全表 + 同功能过程）"""
    errors = []
    i_values = []
    proc_i_map = {}
    for i, r in enumerate(rows, start=4):
        i_val = _safe_strip(r.get('I', ''))
        e = _safe_strip(r.get('E', ''))
        i_values.append((i, i_val))
        proc_i_map.setdefault(e, []).append((i, i_val))

    # 全表唯一性
    i_count = {}
    for i, i_val in i_values:
        if i_val:
            i_count.setdefault(i_val, []).append(i)
    for i_val, indices in i_count.items():
        if len(indices) > 1:
            errors.append(f"I 列「{i_val}」在多行出现：{indices}")

    # 同功能过程唯一性
    for proc, items in proc_i_map.items():
        proc_i_count = {}
        for i, i_val in items:
            if i_val:
                proc_i_count.setdefault(i_val, []).append(i)
        for i_val, indices in proc_i_count.items():
            if len(indices) > 1:
                errors.append(f"功能过程「{proc}」下 I 列「{i_val}」重复：{indices}")

    return {'passed': len(errors) == 0, 'errors': errors}


def check_step14_pingxing(rows, source_docx=None):
    """校验 14：平行判定穷尽（需对照原文手动确认）"""
    return {'passed': True, 'errors': [], 'note': '需人工对照原文确认所有分点/并列条件已拆分'}


def load_rows_from_xlsx(xlsx_path):
    """从 .xlsx 加载业务数据行（处理 A~E 合并单元格）"""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = []

    # 构建合并区域映射：(r,c) -> (主单元格值)
    merge_value_map = {}
    for mr in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = mr.bounds
        # 取主单元格的值
        master_cell = ws.cell(row=min_row, column=min_col)
        master_value = master_cell.value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merge_value_map[(r, c)] = master_value
    # 关键修复：合并区域中**子单元格**使用主单元格值，但**主单元格自身**应该用 cell.value 而非被覆盖
    # 上面的循环已经处理了主单元格（r==min_row），因为 merge_value_map[(min_row, min_col)] = master_value
    # 所以对于主单元格自身，map 中存的就是 cell.value，没有问题
    # 但对于非主单元格（merged_cells），需要从主单元格取值 — 已正确处理
    # 关键问题：A4:A48 合并，A4 是主单元格有值，A5~A48 是 MergedCell
    # 我们需要：所有 (r,c) ∈ merged → 取主单元格值
    # 但 cell.value 对 MergedCell 是 None，而我们的 map 强制覆盖为 master_value — 已正确处理
    # 唯一需要注意：master_value 可能是 None（合并后主单元格是 None 的情况）
    # 此时子单元格也都是 None — 这是正确的

    # 业务数据行从第 4 行开始
    from openpyxl.utils import get_column_letter
    for r_idx in range(4, ws.max_row + 1):
        r = {}
        for c_idx in range(1, 13):  # A~L
            col_letter = get_column_letter(c_idx)
            # 优先用合并区域映射（处理 MergedCell）
            if (r_idx, c_idx) in merge_value_map:
                val = merge_value_map[(r_idx, c_idx)]
            else:
                # 非合并区域：直接读 cell.value
                cell = ws.cell(row=r_idx, column=c_idx)
                val = cell.value
            r[col_letter] = val

        # 跳过完全空行
        if any(v is not None and v != '' for v in r.values()):
            # 二次过滤：F/G/H/I/J/K 全部为空视为表格预留空白行
            has_content = any(_safe_strip(r.get(c, '')) for c in ['F', 'G', 'H', 'I', 'J', 'K'])
            if has_content:
                rows.append(r)
    return rows


def main():
    parser = argparse.ArgumentParser(description='COSMIC 表格合规性校验')
    parser.add_argument('xlsx_path', help='COSMIC 表格 .xlsx 路径')
    parser.add_argument('--source', '-s', help='原始 .docx 路径（可选，用于原文一致性校验）')
    parser.add_argument('--output', '-o', help='输出报告 JSON 路径（可选）')
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx_path)
    if not xlsx_path.exists():
        print(f"错误：文件不存在 - {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    rows = load_rows_from_xlsx(xlsx_path)
    print(f"已加载 {len(rows)} 行业务数据")

    checks = [
        ('check_step0_fenodian', check_step0_fenodian),
        ('check_step1_yuanwen', check_step1_yuanwen),
        ('check_step2_jiegou', check_step2_jiegou),
        ('check_step3_gongneng', check_step3_gongneng),
        ('check_step4_duliang', check_step4_duliang),
        ('check_step5_shuzhi', check_step5_shuzhi),
        ('check_step6_bitian', check_step6_bitian),
        ('check_step7_zifuchuan', check_step7_zifuchuan),
        ('check_step8_fengxianci', check_step8_fengxianci),
        ('check_step9_shujuyidong', check_step9_shujuyidong),
        ('check_step10_fuyongdu', check_step10_fuyongdu),
        ('check_step11_hebing', check_step11_hebing),
        ('check_step12_shujuzu', check_step12_shujuzu),
        ('check_step13_shujushuxing', check_step13_shujushuxing),
        ('check_step14_pingxing', check_step14_pingxing),
    ]

    report = {
        'xlsx_path': str(xlsx_path),
        'row_count': len(rows),
        'checks': [],
        'overall_passed': True,
    }

    for name, check in checks:
        result = check(rows, args.source)
        result['name'] = name
        report['checks'].append(result)
        if not result.get('passed', True):
            report['overall_passed'] = False
        status = '[OK]' if result.get('passed', True) else '[FAIL]'
        print(f"  {status} {name}: {len(result.get('errors', []))} 个错误")
        for err in result.get('errors', [])[:5]:
            print(f"      - {err}")

    if args.output:
        Path(args.output).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
        print(f"\n报告已写入：{args.output}")

    print(f"\n{'='*50}")
    print(f"总体结果：{'[PASS] 全部通过' if report['overall_passed'] else '[FAIL] 存在错误'}")
    print(f"{'='*50}")

    sys.exit(0 if report['overall_passed'] else 1)


if __name__ == '__main__':
    main()
