#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
generate_cosmic.py
端到端从 .docx 概要设计说明书生成 COSMIC 功能点表格 .xlsx
用法：python generate_cosmic.py <docx_path> <output_xlsx_path>
"""
import sys
import argparse
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 数据移动类型 G 列判定
KEYWORDS_E = ['解析入参', '获取入参', '接收推送', '调用接口入参', '获取入参', '入参获取', '接口入参']
KEYWORDS_R = ['查询', '读取', '解析', '校验', '判断']
KEYWORDS_W = ['保存', '写入', '更新', '记录', '生成']
KEYWORDS_X = ['反馈', '返回', '推送', '发送', '调用.*接口', '触发发送']

# 复用度配比（新增:复用:利旧 = 2:7:1）
REUSE_RATIO = {'新增': 0.20, '复用': 0.70, '利旧': 0.10}
CFP_VALUE = {'新增': 1.00, '复用': 0.33, '利旧': 0.00}


def classify_data_movement(desc):
    """根据子过程描述判定数据移动类型"""
    for kw in KEYWORDS_E:
        if kw in desc:
            return 'E'
    for kw in KEYWORDS_R:
        if kw in desc and not any(w in desc for w in ['保存', '写入', '更新']):
            return 'R'
    for kw in KEYWORDS_W:
        if kw in desc:
            return 'W'
    for kw in KEYWORDS_X:
        if kw in desc:
            return 'X'
    return 'R'  # 默认 R


def allocate_reuse(n):
    """按 2:7:1 分配复用度行号"""
    new_count = max(int(round(n * REUSE_RATIO['新增'])), 1)
    old_count = max(int(round(n * REUSE_RATIO['利旧'])), 1)
    reuse_count = n - new_count - old_count

    # 调整到 20%~25% 区间
    if new_count < n * 0.20:
        new_count = int(n * 0.20) + 1
        old_count = max(int(round(n * REUSE_RATIO['利旧'])), 1)
        reuse_count = n - new_count - old_count

    return ['新增'] * new_count + ['复用'] * reuse_count + ['利旧'] * old_count


def write_xlsx(output_path, data_rows, level1_name):
    """写入 .xlsx 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = '2、功能点拆分表'

    # 边框样式
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 第 1 行：12 个空单元格
    for col in range(1, 13):
        ws.cell(row=1, column=col, value='').border = border

    # 第 2 行：合并标题
    ws.cell(row=2, column=1, value='度量策略阶段')
    ws.merge_cells('A2:C2')
    ws.cell(row=2, column=4, value='映射阶段')
    ws.merge_cells('D2:J2')
    ws.cell(row=2, column=11, value='度量阶段')
    ws.merge_cells('K2:L2')
    for col in range(1, 13):
        c = ws.cell(row=2, column=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        c.border = border

    # 第 3 行：表头
    headers = ['客户需求', '功能用户', '功能用户需求', '触发事件', '功能过程',
               '子过程描述', '数据移动类型', '数据组', '数据属性',
               '复用度', 'CFP', 'ΣCFP']
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        c.border = border

    # 列宽
    widths = [30, 22, 30, 12, 30, 32, 8, 30, 32, 8, 8, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 业务数据行
    sum_cfp = 0.0
    for ri, row_data in enumerate(data_rows, start=4):
        for ci, key in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'], start=1):
            val = row_data.get(key, '')
            if ci == 1:
                val = level1_name  # A 列强制 = 核心需求全称
            if ci == 12 and ri != 4:
                val = ''  # L 列仅首行
            if ci == 12 and ri == 4:
                val = ''  # 下面统一计算
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            c.border = border
        sum_cfp += float(row_data.get('K', 0))

    # 填 ΣCFP（仅首行）
    if data_rows:
        ws.cell(row=4, column=12, value=round(sum_cfp, 2))

    wb.save(str(output_path))
    return sum_cfp


def main():
    parser = argparse.ArgumentParser(description='从 .docx 生成 COSMIC 表格')
    parser.add_argument('docx_path', help='概要设计说明书 .docx 路径')
    parser.add_argument('output_path', help='输出 .xlsx 路径')
    parser.add_argument('--intermediate', help='中间提取结果 JSON 路径（可选）')
    args = parser.parse_args()

    # 此脚本为骨架实现，需要配合 extract_requirements.py 提取后人工标注
    # 完整端到端实现需要 LLM 辅助拆分子过程，超出纯脚本能力
    print("=" * 60)
    print("COSMIC 端到端生成器（v1.0 骨架）")
    print("=" * 60)
    print()
    print("说明：")
    print("  本脚本提供 .xlsx 写入、合并、ΣCFP 计算、复用度配比等")
    print("  基础设施。子过程拆分（F/G/H/I 列填充）需要 LLM 辅助，")
    print("  建议流程：")
    print("    1) python scripts/extract_requirements.py <docx> -o extracted.json")
    print("    2) LLM 读取 extracted.json，按 cosmic-rules.md 拆分子过程")
    print("    3) 将拆解结果按本脚本 data_rows 格式传入 write_xlsx()")
    print("    4) python scripts/validate_cosmic.py <output.xlsx> -s <docx>")
    print()
    print("已完成：脚本基础设施就绪")
    print(f"输入文件：{args.docx_path}")
    print(f"输出文件：{args.output_path}")
    print("=" * 60)


if __name__ == '__main__':
    main()
