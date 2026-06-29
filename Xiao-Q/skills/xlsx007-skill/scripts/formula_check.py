#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""formula_check.py - Static validation of a workbook's formulas & cached values.

Usage:
    python3 formula_check.py file.xlsx --report
    python3 formula_check.py file.xlsx --json

Exit code 0 = no errors found; 1 = problems found; 2 = usage error.
Flags Excel error tokens in cached values and in formula text. Does not execute
formulas - run libreoffice_recalc.py first if you need fresh cached values.
"""
import argparse
import json
import sys

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except Exception as e:  # noqa: BLE001
    print("openpyxl is required: " + str(e), file=sys.stderr)
    sys.exit(2)

ERROR_TOKENS = ["#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"]


def _scan(path):
    issues = []
    # 1) cached values
    wb = load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        for r, row in enumerate(ws.iter_rows(), start=1):
            for c, cell in enumerate(row, start=1):
                v = cell.value
                if isinstance(v, str):
                    for tok in ERROR_TOKENS:
                        if tok in v:
                            issues.append({"sheet": ws.title,
                                           "cell": get_column_letter(c) + str(r),
                                           "type": "cached_error", "detail": v})
                            break
    wb.close()
    # 2) formula text
    wb = load_workbook(path, data_only=False, read_only=True)
    for ws in wb.worksheets:
        for r, row in enumerate(ws.iter_rows(), start=1):
            for c, cell in enumerate(row, start=1):
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    for tok in ERROR_TOKENS:
                        if tok in v:
                            issues.append({"sheet": ws.title,
                                           "cell": get_column_letter(c) + str(r),
                                           "type": "formula_error", "detail": v})
                            break
    wb.close()
    return issues


def main():
    ap = argparse.ArgumentParser(description="Static formula / error-token check")
    ap.add_argument("file")
    ap.add_argument("--json", action="store_true")
    ap.add_argument("--report", action="store_true")
    args = ap.parse_args()

    issues = _scan(args.file)
    ok = len(issues) == 0
    if args.json:
        print(json.dumps({"ok": ok, "issues": issues}, ensure_ascii=False, indent=2))
    else:
        if ok:
            print("OK: no error tokens found in " + args.file)
        else:
            print("FOUND " + str(len(issues)) + " issue(s) in " + args.file + ":")
            for i in issues:
                print("  [" + i["type"] + "] " + i["sheet"] + "!" + i["cell"]
                      + " -> " + str(i["detail"]))
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
