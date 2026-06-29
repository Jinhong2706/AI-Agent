#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""xlsx_reader.py - Structure discovery + structural diff for the XLSX skill.

This is a self-contained superset of a basic reader: it adds --diff-against
for verifying that an EDIT did not damage the workbook (sheets, named ranges,
macros, pivots, charts, and a sample of cell values).

It is READ-ONLY: it never modifies any file.

Usage:
    python3 xlsx_reader.py FILE                       # overview
    python3 xlsx_reader.py FILE --sheet NAME --preview 20
    python3 xlsx_reader.py FILE --json
    python3 xlsx_reader.py NEW.xlsx --diff-against OLD.xlsx

Works with .xlsx / .xlsm / .xltx. CSV/TSV: use pandas directly (see SKILL.md
section 10 for Chinese encoding handling).
"""
import argparse
import json
import sys
import zipfile

try:
    from openpyxl import load_workbook
except Exception as e:  # noqa: BLE001
    print(f"openpyxl is required: {e}", file=sys.stderr)
    sys.exit(2)


def _zip_features(path):
    """Detect special parts directly from the package (robust, no parsing)."""
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
    return {
        "has_vba_macros": "xl/vbaProject.bin" in names,
        "has_pivot_tables": any("pivotTable" in n for n in names),
        "has_pivot_cache": any("pivotCache" in n for n in names),
        "has_charts": any(n.startswith("xl/charts/") for n in names),
        "has_shared_strings": "xl/sharedStrings.xml" in names,
        "has_calc_chain": "xl/calcChain.xml" in names,
        "worksheet_parts": sorted(n for n in names if n.startswith("xl/worksheets/sheet")),
    }


def describe(path):
    info = {"file": path}
    info.update(_zip_features(path))
    wb = load_workbook(path, read_only=True, keep_vba=True, data_only=False)
    info["sheets"] = []
    for ws in wb.worksheets:
        info["sheets"].append({
            "name": ws.title,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "state": ws.sheet_state,
        })
    try:
        info["named_ranges"] = sorted(list(wb.defined_names))
    except Exception:  # noqa: BLE001
        info["named_ranges"] = sorted(list(getattr(wb, "defined_names", {}).keys()))
    wb.close()
    return info


def preview(path, sheet, n):
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"sheet not found: {sheet}; available: {wb.sheetnames}")
    ws = wb[sheet]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= n:
            break
        rows.append(["" if v is None else v for v in row])
    wb.close()
    return rows


def _sheet_sample(path, max_sheets=50, max_cells=200):
    """Return {sheet: {cellref: value}} sample for diffing."""
    wb = load_workbook(path, read_only=True, data_only=False)
    sample = {}
    from openpyxl.utils import get_column_letter
    for ws in wb.worksheets[:max_sheets]:
        cells = {}
        count = 0
        for r, row in enumerate(ws.iter_rows(), start=1):
            for c, cell in enumerate(row, start=1):
                if cell.value is None:
                    continue
                cells[f"{get_column_letter(c)}{r}"] = cell.value
                count += 1
                if count >= max_cells:
                    break
            if count >= max_cells:
                break
        sample[ws.title] = cells
    wb.close()
    return sample


def diff_against(new_path, old_path):
    a = describe(old_path)
    b = describe(new_path)
    report = {"old": old_path, "new": new_path, "ok": True, "issues": [], "notes": []}

    old_sheets = [s["name"] for s in a["sheets"]]
    new_sheets = [s["name"] for s in b["sheets"]]
    if old_sheets != new_sheets:
        report["ok"] = False
        report["issues"].append(
            f"sheet set/order changed: {old_sheets} -> {new_sheets}")

    for feat in ("has_vba_macros", "has_pivot_tables", "has_charts"):
        if a[feat] and not b[feat]:
            report["ok"] = False
            report["issues"].append(f"LOST {feat}: present in old, missing in new")

    if set(a["named_ranges"]) - set(b["named_ranges"]):
        missing = sorted(set(a["named_ranges"]) - set(b["named_ranges"]))
        report["ok"] = False
        report["issues"].append(f"named ranges removed: {missing}")

    # Cell-level changes on shared sheets (best-effort sample).
    try:
        sa = _sheet_sample(old_path)
        sb = _sheet_sample(new_path)
        changes = []
        for sheet in set(sa) & set(sb):
            for ref, ov in sa[sheet].items():
                nv = sb[sheet].get(ref, "<removed>")
                if nv != ov:
                    changes.append(f"{sheet}!{ref}: {ov!r} -> {nv!r}")
        report["changed_cells_sample"] = changes[:100]
        report["notes"].append(f"{len(changes)} changed cell(s) in sample")
    except Exception as e:  # noqa: BLE001
        report["notes"].append(f"cell sample diff skipped: {e}")

    return report


def _print_overview(info):
    print(f"File: {info['file']}")
    print(f"Sheets ({len(info['sheets'])}):")
    for s in info["sheets"]:
        print(f"  - {s['name']}: {s['rows']} rows x {s['cols']} cols [{s['state']}]")
    print(f"Named ranges: {info['named_ranges'] or '(none)'}")
    print(f"VBA macros : {info['has_vba_macros']}")
    print(f"Pivot tables: {info['has_pivot_tables']}")
    print(f"Charts     : {info['has_charts']}")
    print(f"calcChain  : {info['has_calc_chain']}")


def main():
    ap = argparse.ArgumentParser(description="XLSX structure reader / differ (read-only)")
    ap.add_argument("file")
    ap.add_argument("--sheet")
    ap.add_argument("--preview", type=int, default=0)
    ap.add_argument("--diff-against", dest="diff_against", metavar="OLD.xlsx")
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()

    if args.diff_against:
        rep = diff_against(args.file, args.diff_against)
        if args.json:
            print(json.dumps(rep, ensure_ascii=False, indent=2, default=str))
        else:
            print(f"DIFF  old={rep['old']}  new={rep['new']}")
            print("RESULT:", "OK (no structural damage detected)" if rep["ok"] else "PROBLEMS FOUND")
            for i in rep["issues"]:
                print("  ! " + i)
            for n in rep["notes"]:
                print("  . " + n)
            for c in rep.get("changed_cells_sample", [])[:20]:
                print("    ~ " + c)
        return 0 if rep["ok"] else 1

    info = describe(args.file)
    if args.preview and args.sheet:
        rows = preview(args.file, args.sheet, args.preview)
        if args.json:
            print(json.dumps({"overview": info, "preview": rows}, ensure_ascii=False, indent=2, default=str))
        else:
            _print_overview(info)
            print(f"\nPreview of '{args.sheet}' (first {args.preview} rows):")
            for row in rows:
                print(" | ".join(str(v) for v in row))
        return 0

    if args.json:
        print(json.dumps(info, ensure_ascii=False, indent=2, default=str))
    else:
        _print_overview(info)
    return 0


if __name__ == "__main__":
    sys.exit(main())
